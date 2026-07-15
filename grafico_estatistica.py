# -*- coding: utf-8 -*-
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from scipy import stats as spstats

sns.set_style('whitegrid')

wb = load_workbook('coleta_ampliada.xlsx', data_only=True)
sheets_config = {
    'Produtos': range(6, 14),
    'Payment Methods': range(6, 12),
    'Customers': range(6, 12),
    'Shipping Methods': range(6, 10),
}
rows = []
for sheet_name, row_range in sheets_config.items():
    ws = wb[sheet_name]
    for r in row_range:
        smells = ws[f'E{r}'].value
        if smells is None:
            continue
        deteccao = ws[f'D{r}'].value or 0
        redundancia = ws[f'F{r}'].value or 0
        rows.append({'Detecção': deteccao, 'Test smells': smells, 'Redundância': redundancia})
df = pd.DataFrame(rows)

fig, axes = plt.subplots(1, 2, figsize=(11, 4.3))

# Painel (a): heatmap de correlação de Spearman
corr = df.corr(method='spearman')
sns.heatmap(corr, annot=True, fmt='.2f', cmap='RdBu_r', vmin=-1, vmax=1,
            square=True, ax=axes[0], cbar_kws={'label': 'rho (Spearman)'})
axes[0].set_title('(a) Correlação de Spearman entre métricas', fontweight='bold', fontsize=10)

# Painel (b): média +/- IC95% (bootstrap)
rng = np.random.default_rng(42)
metrics = ['Detecção', 'Test smells', 'Redundância']
labels = ['Detecção (%)', 'Test smells (0-7)', 'Redundância (%)']
means, los, his = [], [], []
for m in metrics:
    data = (df[m].to_numpy(),)
    res = spstats.bootstrap(data, np.mean, n_resamples=10000, random_state=rng, method='percentile')
    means.append(df[m].mean())
    los.append(res.confidence_interval.low)
    his.append(res.confidence_interval.high)

colors = ['#4C72B0', '#55A868', '#C44E52']
for i, (m, mean, lo, hi, c) in enumerate(zip(metrics, means, los, his, colors)):
    scale = 1 if m == 'Test smells' else 100
    val, lo_v, hi_v = mean * scale, lo * scale, hi * scale
    axes[1].errorbar(i, val, yerr=[[val - lo_v], [hi_v - val]], fmt='o', color=c,
                     markersize=9, capsize=6, capthick=2, elinewidth=2)

axes[1].set_xticks([0, 1, 2])
axes[1].set_xticklabels(labels, fontsize=8.5)
axes[1].set_xlim(-0.5, 2.5)
axes[1].set_ylabel('Média (% ou nº de tipos)', fontsize=9)
axes[1].set_title('(b) Média e IC 95% (bootstrap)', fontweight='bold', fontsize=10)
axes[1].grid(axis='y', linestyle=':', alpha=0.6)
axes[1].set_axisbelow(True)

plt.tight_layout()
plt.savefig('estatistica_figuras.pdf', bbox_inches='tight')
plt.savefig('estatistica_figuras.png', dpi=200, bbox_inches='tight')
print("Figura gerada com sucesso!")
