import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import load_workbook

# ============================================================
# 1. Carregar os dados diretamente da planilha de coleta
# ============================================================
wb = load_workbook('../data/coleta_ampliada.xlsx', data_only=True)

sheets_config = {
    'Produtos': range(6, 14),
    'Payment Methods': range(6, 12),
    'Customers': range(6, 12),
    'Shipping Methods': range(6, 10),
}

rows = []
for sheet_name, row_range in sheets_config.items():
    ws = wb[sheet_name]
    for r in row_range:
        smells = ws[f'E{r}'].value
        if smells is None:
            continue  # linha excluida (duplicata de dupla ou nao-entrega)
        deteccao = ws[f'D{r}'].value or 0
        redundancia = ws[f'F{r}'].value or 0
        rows.append({
            'Funcionalidade': sheet_name,
            'Deteccao': deteccao,
            'Smells': smells,
            'Redundancia': redundancia,
        })

df = pd.DataFrame(rows)
print(f"Total de suites carregadas: {len(df)}")
print(df.groupby('Funcionalidade').size().to_string())
print()

# ============================================================
# 2. Teste de normalidade (Shapiro-Wilk) - amostra completa N=22
# ============================================================
print("=" * 60)
print("TESTE DE NORMALIDADE (Shapiro-Wilk, N=22)")
print("=" * 60)
for col in ['Deteccao', 'Smells', 'Redundancia']:
    stat, p = stats.shapiro(df[col])
    normal = "SIM" if p > 0.05 else "NAO"
    print(f"{col:15s}: W={stat:.4f}, p={p:.6f} -> Normal? {normal}")
print()

# ============================================================
# 3. Teste de hipotese: suites discentes vs. valor fixo da referencia
#    (Wilcoxon signed-rank, one-sample)
# ============================================================
print("=" * 60)
print("TESTE DE HIPOTESE: Discentes vs. Referencia (Wilcoxon)")
print("=" * 60)

referencia = {'Deteccao': 1.0, 'Smells': 0, 'Redundancia': 0.0}

for col, ref_val in referencia.items():
    diffs = df[col] - ref_val
    non_zero = diffs[diffs != 0]
    if len(non_zero) < 1:
        print(f"{col:15s}: sem variacao suficiente para o teste")
        continue
    try:
        stat, p = stats.wilcoxon(diffs)
        sig = "SIGNIFICATIVO (p<0.05)" if p < 0.05 else "nao significativo"
        print(f"{col:15s}: W={stat:.2f}, p={p:.6f} -> {sig}")
        print(f"                 (n={len(df)}, diferencas nao-nulas={len(non_zero)})")
    except Exception as e:
        print(f"{col:15s}: erro no teste - {e}")
print()

# ============================================================
# 4. Correlacao de Spearman entre as 3 metricas
# ============================================================
print("=" * 60)
print("CORRELACAO DE SPEARMAN entre metricas (N=22)")
print("=" * 60)
pairs = [('Deteccao', 'Smells'), ('Deteccao', 'Redundancia'), ('Smells', 'Redundancia')]
for a, b in pairs:
    rho, p = stats.spearmanr(df[a], df[b])
    sig = "significativa (p<0.05)" if p < 0.05 else "nao significativa"
    print(f"{a} x {b}: rho={rho:.3f}, p={p:.4f} -> {sig}")
print()

# ============================================================
# 5. Intervalo de confianca (95%) via bootstrap para as medias
# ============================================================
print("=" * 60)
print("INTERVALO DE CONFIANCA 95% (bootstrap, 10000 reamostragens)")
print("=" * 60)
rng = np.random.default_rng(42)
for col in ['Deteccao', 'Smells', 'Redundancia']:
    data = (df[col].to_numpy(),)
    res = stats.bootstrap(data, np.mean, n_resamples=10000, random_state=rng, method='percentile')
    mean = df[col].mean()
    print(f"{col:15s}: media={mean:.4f}, IC95%=[{res.confidence_interval.low:.4f}, {res.confidence_interval.high:.4f}]")
print()

print("Analise concluida. Copie estes resultados para a secao de Resultados do artigo.")
